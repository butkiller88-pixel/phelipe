import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, date
import os

# Configuração da página
st.set_page_config(page_title="Gestão de Obra Unificada", layout="wide", initial_sidebar_state="expanded")

# --- CONFIGURAÇÃO DE CAMINHOS (Adaptar para o ambiente Zapia) ---
# No Zapia, usamos arquivos locais no workspace.
CAMINHO_OBRA = "Controle_de_Obra.xlsx"
CAMINHO_ESTOQUE = "Controle_de_Estoque.xlsx"

# --- ESTILOS CSS ---
st.markdown("""
    <style>
    .main { background-color: #0F172A; color: #F8FAFC; }
    .stTabs [data-baseweb="tab-list"] { gap: 10px; }
    .stTabs [data-baseweb="tab"] {
        background-color: #1E293B;
        border-radius: 8px 8px 0px 0px;
        padding: 10px 20px;
        color: #94A3B8;
    }
    .stTabs [aria-selected="true"] {
        background-color: #3B82F6 !important;
        color: white !important;
    }
    div.stButton > button {
        width: 100%;
        background-color: #10B981;
        color: white;
        border: none;
        padding: 0.5rem;
        font-weight: bold;
    }
    .kpi-card {
        background-color: #1E293B;
        padding: 20px;
        border-radius: 15px;
        border-left: 5px solid #3B82F6;
        margin-bottom: 20px;
    }
    </style>
    """, unsafe_allow_html=True)

# --- FUNÇÕES DE CARREGAMENTO ---
def formatar_moeda(valor):
    try:
        return f"R$ {float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return "R$ 0,00"

def load_data_obra():
    if not os.path.exists(CAMINHO_OBRA):
        return None
    return openpyxl.load_workbook(CAMINHO_OBRA, data_only=True)

def load_data_estoque():
    if not os.path.exists(CAMINHO_ESTOQUE):
        return None
    return openpyxl.load_workbook(CAMINHO_ESTOQUE, data_only=True)

def save_to_excel(file_path, sheet_name, row_data):
    if not os.path.exists(file_path):
        st.error(f"Arquivo {file_path} não encontrado.")
        return False
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]
        
        # Localizar linha do TOTAL
        linha_total_idx = None
        for row in range(5, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=2).value
            if cell_val and str(cell_val).strip().upper().startswith("TOTAL"):
                linha_total_idx = row
                break
        
        if not linha_total_idx:
            linha_total_idx = ws.max_row + 1
            ws.cell(row=linha_total_idx, column=2, value="TOTAL")
            
        ws.insert_rows(linha_total_idx, amount=1)
        
        # Estilos (Simulado)
        for c_idx, val in enumerate(row_data, start=2):
            cell = ws.cell(row=linha_total_idx, column=c_idx, value=val)
            if isinstance(val, (datetime, date)):
                cell.number_format = 'dd/mm/yyyy'
        
        wb.save(file_path)
        return True
    except Exception as e:
        st.error(f"Erro ao salvar: {e}")
        return False

# --- SIDEBAR (NAVEGAÇÃO) ---
st.sidebar.title("🏗️ Zapia Obra")
menu = st.sidebar.radio("Navegação", ["Dashboard Geral", "Financeiro (Lançar)", "Estoque (Lançar)", "Cronograma", "Diário de Obra"])

# --- DASHBOARD GERAL ---
if menu == "Dashboard Geral":
    st.title("📊 Painel de Controle de Obra")
    wb_obra = load_data_obra()
    wb_estoque = load_data_estoque()
    
    if wb_obra:
        # Cálculos Simples baseados no script gerar_relatorios.py
        # Nota: Em produção, leríamos os valores reais das planilhas.
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Total Entradas", "R$ 150.111,00")
        with col2:
            st.metric("Total Saídas", "R$ 11.680,00", delta="-R$ 380,00", delta_color="inverse")
        with col3:
            st.metric("Saldo Atual", "R$ 138.431,00")
        with col4:
            st.metric("Progresso Físico", "52.5%")
            st.progress(0.525)

    if wb_estoque:
        st.divider()
        st.subheader("📦 Alertas de Estoque")
        st.warning("⚠️ **Adesivo PVC:** Saldo 4 (Mínimo: 10)")
        st.error("🚨 **Cimento CP-II:** Saldo 2 (Mínimo: 15)")

# --- FINANCEIRO (LANÇAR) ---
elif menu == "Financeiro (Lançar)":
    st.title("💰 Lançamentos Financeiros")
    tab1, tab2 = st.tabs(["📤 Saída (Despesa)", "📥 Entrada (Aporte)"])
    
    with tab1:
        with st.form("form_saida"):
            data_s = st.date_input("Data", date.today())
            etapa_s = st.selectbox("Etapa", ["01. Serviços Preliminares", "02. Infraestrutura", "03. Supraestrutura", "04. Alvenaria", "05. Cobertura", "06. Instalações", "07. Acabamento"])
            desc_s = st.text_input("Item / Descrição")
            forn_s = st.text_input("Fornecedor")
            val_p_s = st.number_input("Valor Previsto", min_value=0.0)
            val_r_s = st.number_input("Valor Real (Pago)", min_value=0.0)
            forma_s = st.selectbox("Forma de Pagamento", ["Pix", "Boleto", "Cartão", "Dinheiro"])
            status_s = st.selectbox("Status", ["Pago", "Pendente", "Agendado"])
            
            if st.form_submit_button("Gravar Saída"):
                if desc_s:
                    # [data, etapa, desc, forn, val_p, val_r, forma, status]
                    res = save_to_excel(CAMINHO_OBRA, "Custos", [data_s, etapa_s, desc_s, forn_s, val_p_s, val_r_s, forma_s, status_s])
                    if res: st.success("Saída registrada!")
                else: st.warning("Preencha a descrição.")

    with tab2:
        with st.form("form_entrada"):
            data_e = st.date_input("Data ", date.today())
            desc_e = st.text_input("Descrição / Origem")
            val_e = st.number_input("Valor Recebido", min_value=0.0)
            forma_e = st.selectbox("Forma", ["Pix", "Transferência", "Dinheiro"])
            status_e = st.selectbox("Status ", ["Recebido", "Pendente"])
            
            if st.form_submit_button("Gravar Entrada"):
                if desc_e:
                    res = save_to_excel(CAMINHO_OBRA, "Entradas", [data_e, desc_e, val_e, forma_e, status_e])
                    if res: st.success("Entrada registrada!")
                else: st.warning("Preencha a descrição.")

# --- ESTOQUE (LANÇAR) ---
elif menu == "Estoque (Lançar)":
    st.title("📦 Movimentação de Estoque")
    with st.form("form_estoque"):
        tipo_m = st.radio("Tipo de Movimento", ["Entrada", "Saída"], horizontal=True)
        codigo_m = st.text_input("Código do Item")
        nome_m = st.text_input("Nome do Material")
        qtd_m = st.number_input("Quantidade", min_value=0.1)
        resp_m = st.text_input("Responsável")
        dest_m = st.text_input("Destino / Origem")
        
        if st.form_submit_button("Registrar Movimento"):
            if codigo_m and nome_m:
                # [data, codigo, nome, tipo, qtd, un, resp, destino, nf, obs]
                res = save_to_excel(CAMINHO_ESTOQUE, "Movimentacoes", [date.today(), codigo_m, nome_m, tipo_m, qtd_m, "un", resp_m, dest_m, "", "Via App Web"])
                if res: st.success("Estoque atualizado!")
            else: st.warning("Código e Nome são obrigatórios.")

# --- CRONOGRAMA ---
elif menu == "Cronograma":
    st.title("📅 Cronograma da Obra")
    # Exemplo de visualização baseada no script de relatório
    atividades = [
        {"Atividade": "Montagem de Pilares", "Fim": "05/07/2026", "Resp": "Marcos", "Progresso": 85},
        {"Atividade": "Levantamento das Paredes", "Fim": "20/07/2026", "Resp": "Pedro", "Progresso": 40},
    ]
    df = pd.DataFrame(atividades)
    st.table(df)

# --- DIÁRIO DE OBRA ---
elif menu == "Diário de Obra":
    st.title("📝 Diário de Obra")
    with st.form("form_diario"):
        data_d = st.date_input("Data", date.today())
        clima_d = st.selectbox("Clima", ["Ensolarado", "Nublado", "Chuva Leve", "Chuva Forte"])
        equipe_d = st.text_input("Equipe Presente")
        atividades_d = st.text_area("Atividades Realizadas")
        obs_d = st.text_area("Ocorrências / Observações")
        
        if st.form_submit_button("Salvar Diário"):
            res = save_to_excel(CAMINHO_OBRA, "Diario_de_Obra", [data_d, clima_d, equipe_d, atividades_d, obs_d])
            if res: st.success("Diário salvo com sucesso!")

st.sidebar.divider()
st.sidebar.caption("© 2026 Grupo Bom Jardim | Zapia AI")
